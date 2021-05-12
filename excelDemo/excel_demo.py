#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county.
import openpyxl, pprint

print('opening workbook......')
wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countyData = {}
print('max_row:' + str(sheet.max_row))
print('Reading rows......')
for row in range(2, sheet.max_row + 1):
    # 州
    state = sheet['B' + str(row)].value
    # 县
    county = sheet['C' + str(row)].value
    # 人数
    pop = sheet['D' + str(row)].value

    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)

print('writing results....')
with open('census2010.py', 'w') as fp:
    fp.write('allData = ' + pprint.pformat(countyData))

print('Done')

